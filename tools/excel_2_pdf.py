import os
import tempfile
import time
from typing import Any, Dict, List, Generator, Tuple, Optional
import copy

from dify_plugin import Tool
from dify_plugin.entities.tool import ToolInvokeMessage
from dify_plugin.file.file import File

# 导入依赖库，包含错误处理
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape, portrait
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    DEPENDENCIES_AVAILABLE = True
except ImportError:
    DEPENDENCIES_AVAILABLE = False

class ExcelToPdfTool(Tool):
    """
    Excel to PDF Converter with Smart Layout Engine.
    Features: Auto-sizing, Landscape support, Table splitting for wide columns.
    """

    def _invoke(self, tool_parameters: dict[str, Any]) -> Generator[ToolInvokeMessage, None, None]:
        if not DEPENDENCIES_AVAILABLE:
            yield self.create_text_message("Error: Required libraries (openpyxl, reportlab) are missing.")
            return

        input_file = tool_parameters.get("input_file")
        if not input_file:
            yield self.create_text_message("Error: Input file is required.")
            return

        # 1. 验证文件格式
        if not input_file.extension or input_file.extension.lower() not in ['.xlsx']:
            yield self.create_text_message("Error: Only .xlsx files are supported.")
            return

        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                # 2. 准备文件
                input_path = os.path.join(temp_dir, input_file.filename)
                with open(input_path, "wb") as f:
                    f.write(input_file.blob)
                
                output_filename = os.path.splitext(input_file.filename)[0] + ".pdf"
                output_path = os.path.join(temp_dir, output_filename)

                # 3. 执行转换核心逻辑
                converter = ExcelPdfConverter(input_path, output_path)
                result = converter.convert()

                if not result["success"]:
                    yield self.create_text_message(f"Conversion Failed: {result['message']}")
                    return

                # 4. 读取并返回结果
                with open(output_path, 'rb') as f:
                    pdf_content = f.read()

                yield self.create_text_message("Conversion successful with smart layout optimization.")
                
                yield self.create_blob_message(
                    blob=pdf_content,
                    meta={
                        "filename": output_filename,
                        "mime_type": "application/pdf"
                    }
                )

        except Exception as e:
            yield self.create_text_message(f"System Error: {str(e)}")

class ExcelPdfConverter:
    """
    内部转换器类，负责具体的排版算法和PDF生成
    """
    def __init__(self, input_path: str, output_path: str):
        self.input_path = input_path
        self.output_path = output_path
        self.font_name = "CustomChineseFont"
        self.font_bold_name = "CustomChineseFont" # 只有单独字体文件时，粗体也用同一个
        self.registered_font = False
        
        # 页面配置 (A4)
        self.page_width_pt_portrait = A4[0]  # ~595
        self.page_height_pt_portrait = A4[1] # ~842
        self.page_width_pt_landscape = A4[1] 
        self.margin = 30
        
        # 初始化字体
        self._register_fonts()

    def _register_fonts(self):
        """注册自定义字体，路径为 ../fonts/chinese_font.ttc"""
        try:
            # 获取当前脚本所在目录
            current_dir = os.path.dirname(os.path.abspath(__file__))
            # 向上寻找 fonts 目录 (假设结构 plugin/tools/tool.py -> plugin/fonts/)
            # 根据插件结构可能 path 需要调整，这里假设 fonts 与 tools 同级目录或者在根目录下
            # 尝试路径 1: ../fonts/
            font_path = os.path.join(os.path.dirname(current_dir), "fonts", "chinese_font.ttc")
            
            if not os.path.exists(font_path):
                # 备用路径: 当前目录下 fonts/
                font_path = os.path.join(current_dir, "fonts", "chinese_font.ttc")

            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont(self.font_name, font_path))
                self.registered_font = True
            else:
                # 回退：如果找不到字体，使用内置字体（中文会乱码，但至少不报错）
                print(f"Warning: Font file not found at {font_path}, utilizing Helvetica")
                self.font_name = "Helvetica"
                self.font_bold_name = "Helvetica-Bold"
        except Exception as e:
            print(f"Font registration error: {e}")
            self.font_name = "Helvetica"
            self.font_bold_name = "Helvetica-Bold"

    def _clean_cell_text(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _measure_text_width(self, text: str, font_size: int) -> float:
        """精确计算文本宽度的辅助函数"""
        if not text:
            return 0.0
        try:
            return pdfmetrics.stringWidth(text, self.font_name, font_size)
        except:
            return len(text) * font_size * 0.6

    def _get_optimized_columns(self, data: List[List[str]], font_size: int, max_col_width_inch: float = 2.5) -> List[float]:
        """
        计算每列的最佳宽度
        :return: 每一列的宽度列表 (单位: points)
        """
        if not data:
            return []

        num_cols = len(data[0])
        col_widths = [0.0] * num_cols
        
        # 限制最大宽度的点数
        max_width_pts = max_col_width_inch * inch 
        min_width_pts = 20.0 # 最小宽度

        # 采样前100行进行宽度估算（避免数据量过大太慢）
        sample_data = data[:100]

        for row in sample_data:
            for i, cell_text in enumerate(row):
                if i < num_cols:
                    # 增加一些padding
                    w = self._measure_text_width(cell_text, font_size) + 12 
                    if w > col_widths[i]:
                        col_widths[i] = w
        
        # 归一化：应用最大最小值限制
        final_widths = []
        for w in col_widths:
            w = max(min_width_pts, w)
            w = min(max_width_pts, w) # 如果超出最大宽度，后续使用 Paragraph 自动换行
            final_widths.append(w)
            
        return final_widths

    def convert(self) -> Dict[str, Any]:
        try:
            wb = openpyxl.load_workbook(self.input_path, data_only=True)
            story = []
            
            # 使用 ReportLab 的各种样式
            styles = getSampleStyleSheet()
            
            # 定义中文样式
            normal_style = ParagraphStyle(
                name='Normal_CN',
                parent=styles['Normal'],
                fontName=self.font_name,
                fontSize=10,
                leading=12, # 行间距
                alignment=TA_CENTER,
                wordWrap='CJK' # 支持中文换行
            )
            
            title_style = ParagraphStyle(
                name='Title_CN',
                parent=styles['Heading1'],
                fontName=self.font_bold_name,
                fontSize=16,
                leading=20,
                alignment=TA_CENTER,
                spaceAfter=20
            )

            # 页面模板配置
            doc = SimpleDocTemplate(
                self.output_path,
                pagesize=A4,  # 默认初始值，后面会根据内容调整
                leftMargin=self.margin,
                rightMargin=self.margin,
                topMargin=self.margin,
                bottomMargin=self.margin
            )

            # 判断是否需要横向页面
            use_landscape = False

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 提取数据
                raw_data = []
                for row in ws.iter_rows(values_only=True):
                    cleaned_row = [self._clean_cell_text(cell) for cell in row]
                    # 跳过全空行
                    if any(cleaned_row):
                        raw_data.append(cleaned_row)
                
                if not raw_data:
                    continue

                # 添加标题
                story.append(Paragraph(f"Sheet: {sheet_name}", title_style))
                story.append(Spacer(1, 10))

                # --- 核心：智能排版与切分逻辑 ---
                
                # 1. 初步计算列宽 (基于默认10号字体)
                base_font_size = 10
                col_widths = self._get_optimized_columns(raw_data, base_font_size)
                total_width = sum(col_widths)
                
                # 定义可用宽度
                avail_width_portrait = self.page_width_pt_portrait - (2 * self.margin)
                avail_width_landscape = self.page_width_pt_landscape - (2 * self.margin)

                # 决策
                current_data_font_size = base_font_size
                split_tables = False # 是否需要切分表格
                
                if total_width <= avail_width_portrait:
                    # 方案A: 纵向足够
                    pass 
                elif total_width <= avail_width_landscape:
                    # 方案B: 横向足够
                    use_landscape = True
                elif total_width <= avail_width_landscape * 1.25:
                    # 方案C: 横向 + 缩小字体 (最多接受超出25%，通过缩小字体适配)
                    use_landscape = True
                    scale_factor = avail_width_landscape / total_width
                    # 调整列宽和字号
                    col_widths = [w * scale_factor for w in col_widths]
                    current_data_font_size = max(6, int(base_font_size * scale_factor)) # 最小6号字
                else:
                    # 方案D: 表格太宽了，必须切分 (Vertical Slicing)
                    use_landscape = True
                    split_tables = True

                # 构建表格数据 (将普通文本转换为支持换行的 Paragraph)
                # 如果需要切分，逻辑会复杂一些
                
                if not split_tables:
                    # 正常生成一个表格
                    table_data = self._build_table_paragraphs(raw_data, normal_style, current_data_font_size)
                    self._create_and_append_table(story, table_data, col_widths, current_data_font_size)
                else:
                    # 执行切分逻辑
                    self._process_split_tables(story, raw_data, col_widths, avail_width_landscape, normal_style, base_font_size)

                story.append(PageBreak())

            # 设置最终文档页面方向
            if use_landscape:
                doc.pagesize = landscape(A4)
            else:
                doc.pagesize = A4
                
            doc.build(story)
            
            return {"success": True, "message": "PDF created"}

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {"success": False, "message": str(e)}

    def _build_table_paragraphs(self, data, base_style, font_size):
        """将文本数据转换为 ReportLab 的 Paragraph 对象"""
        processed_data = []
        
        # 创建特定字号的样式
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=base_style,
            fontSize=font_size,
            leading=font_size * 1.2
        )

        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=cell_style,
            fontName=self.font_bold_name,
            textColor=colors.whitesmoke
        )

        for row_idx, row in enumerate(data):
            new_row = []
            for cell_val in row:
                style = header_style if row_idx == 0 else cell_style
                new_row.append(Paragraph(cell_val, style))
            processed_data.append(new_row)
        return processed_data

    def _create_and_append_table(self, story, table_data, col_widths, font_size, table_title_suffix=""):
        """创建并添加表格到 story"""
        if table_title_suffix:
             story.append(Paragraph(table_title_suffix, ParagraphStyle('sub', fontSize=8, textColor=colors.grey)))

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.6)), # 深蓝表头
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'), # 顶部对齐以适应换行
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, -1), self.font_name),
            ('FONTSIZE', (0, 0), (-1, -1), font_size),
            ('topPadding', (0, 0), (-1, -1), 4),
            ('bottomPadding', (0, 0), (-1, -1), 4),
        ]
        
        table.setStyle(TableStyle(style_cmds))
        story.append(table)
        story.append(Spacer(1, 20))

    def _process_split_tables(self, story, raw_data, all_col_widths, max_page_width, base_style, font_size):
        """
        切分超宽表格算法
        """
        header_row = raw_data[0]
        data_rows = raw_data[1:]
        
        # 分组逻辑
        slices = [] # 存储 [(start_col_idx, end_col_idx, current_slice_widths)]
        
        current_slice_start = 0
        current_slice_width = 0
        current_slice_widths = []
        
        for i, width in enumerate(all_col_widths):
            # 如果单列就超过了页面宽度，强制将其限制为页面宽度
            if width > max_page_width:
                width = max_page_width
            
            if current_slice_width + width > max_page_width:
                # 当前这一片装不下了，结束当前切片
                slices.append((current_slice_start, i, current_slice_widths))
                # 开启新切片
                current_slice_start = i
                current_slice_width = width
                current_slice_widths = [width]
            else:
                current_slice_width += width
                current_slice_widths.append(width)
        
        # 添加最后一个切片
        if current_slice_widths:
            slices.append((current_slice_start, len(all_col_widths), current_slice_widths))

        # 为每个切片生成表格
        total_parts = len(slices)
        for idx, (start, end, widths) in enumerate(slices):
            # 构建该切片的数据
            # 必须包含表头
            slice_data_raw = []
            
            # 1. 表头部分
            header_slice = header_row[start:end]
            slice_data_raw.append(header_slice)
            
            # 2. 数据部分
            for row in data_rows:
                slice_data_raw.append(row[start:end])
            
            # 构建 Paragraphs
            table_data = self._build_table_paragraphs(slice_data_raw, base_style, font_size)
            
            suffix = f"(Part {idx + 1} of {total_parts})"
            self._create_and_append_table(story, table_data, widths, font_size, table_title_suffix=suffix)